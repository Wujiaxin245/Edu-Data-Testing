import pytest
import pandas as pd
from edu_tests import check_education_data, save_with_stats, load_excel, print_report
from unittest.mock import patch
from openpyxl import load_workbook
import os

# ✅ 原有逻辑：数据校验核心测试
def test_check_education_data():
    df = pd.DataFrame({
        '学号': ['001', '001', '003'],
        '学习时间': pd.to_datetime(['2023-01-01 02:00', '2023-01-01 08:00', '2023-01-01 04:00']),
        '学习时长': [20, 60, 85],
        '学习状态': ['正常', '正常', '正常'],
        '完成状态': ['未完成', '完成', '未完成']
    })

    df_checked, stats = check_education_data(df)

    assert len(df_checked) == 3
    assert '检测结果' in df_checked.columns
    assert stats["总记录数"] == 3
    assert stats["学号重复"] == 1
    assert stats["凌晨学习"] == 2


# ✅ 测试 Excel 写入函数 save_with_stats
def test_save_with_stats(tmp_path):
    df = pd.DataFrame({
        '姓名': ['张三', '李四'],
        '成绩': [90, 85]
    })
    stats = {'正常': 2, '异常': 0}

    output_file = tmp_path / "output.xlsx"
    save_with_stats(df, stats, output_file)

    assert output_file.exists()

    wb = load_workbook(output_file)
    sheets = wb.sheetnames
    assert "数据检测结果" in sheets
    assert "统计报告" in sheets


# ✅ 异常处理：文件不存在时的 load_excel
def test_load_excel_file_not_found():
    with pytest.raises(FileNotFoundError):
        load_excel("this_file_does_not_exist.xlsx")


# ✅ 测试 print_report 函数是否调用 print（mock）
def test_print_report_prints():
    stats = {'正常': 10, '异常': 1}
    with patch('builtins.print') as mock_print:
        print_report(stats)
        assert mock_print.called
        assert any("正常" in str(call.args[0]) for call in mock_print.call_args_list)
